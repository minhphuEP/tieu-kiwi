"""Write TestCase drafts to an Excel workbook following the Testomat.io import
format documented in kb/_global/QE/templates/testcase_template.md. Mirrors the
column conventions scripts/ingest/testcases.py reads, but is not a strict
round-trip inverse — re-ingesting an exported data-driven sheet is not
currently supported by that script's parser.

Interface:
  export_excel(testcases) -> bytes   # .xlsx workbook, ready to upload

Draft schema per testcase (see tieukiwi/testcase_gen.py):
  {"ref", "ac_refs", "title", "type", "priority", "precondition",
   "steps": [...], "data_variants": [...], "api": {...}}
Routed by `type`: "Normal" -> shared Normal_TestCases sheet; "API" -> shared
API_TestCases sheet; "DataTable" -> its own sheet named after `ref`, where each
variant's `values` dict may include an "Expected" key for the per-row expected
result column (kept last).
"""
import io

import openpyxl

_NORMAL_SHEET = "Normal_TestCases"
_NORMAL_HEADERS = ["Title", "Priority", "Pre-condition", "Step_Description",
                   "Test_Data", "Step_ExpectedResult"]
_API_SHEET = "API_TestCases"
_API_HEADERS = ["Title", "Priority", "Pre-condition", "Endpoint", "Method",
                "Request_Headers", "Request_Body", "Expected_Status", "Expected_Response"]
_DATA_SEP_TEXT = "DATA TABLE  ▼  one row = one set of test data"
_INVALID_SHEET_CHARS = set('/\\?*[]:')


def _safe_sheet_name(ref, existing_names):
    name = "".join("_" if ch in _INVALID_SHEET_CHARS else ch for ch in ref)[:31]
    if name not in existing_names:
        return name
    # De-duplicate by truncating further to make room for a numeric suffix.
    for i in range(2, 1000):
        suffix = f"_{i}"
        candidate = (name[: 31 - len(suffix)] + suffix)
        if candidate not in existing_names:
            return candidate
    raise ValueError(f"could not generate a unique sheet name for ref={ref!r}")


def export_excel(testcases):
    """testcases: list of draft-schema dicts. Returns the .xlsx workbook as bytes."""
    if not testcases:
        raise ValueError("export_excel: testcases list is empty")
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # drop the default blank sheet

    api_tcs = [tc for tc in testcases if tc.get("type") == "API"]
    data_driven_tcs = [tc for tc in testcases if tc.get("type") == "DataTable"]
    normal_tcs = [tc for tc in testcases if tc.get("type") not in ("API", "DataTable")]

    if normal_tcs:
        _write_normal_sheet(wb, normal_tcs)
    if api_tcs:
        _write_api_sheet(wb, api_tcs)
    for tc in data_driven_tcs:
        _write_data_driven_sheet(wb, tc)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _write_normal_sheet(wb, testcases):
    ws = wb.create_sheet(_NORMAL_SHEET)
    ws.append(_NORMAL_HEADERS)
    for tc in testcases:
        steps = tc["steps"] or [{"description": "", "expected": ""}]
        first = steps[0]
        ws.append([tc["title"], tc["priority"], tc.get("precondition", ""),
                   first["description"], "(single-action TC — no DT)", first["expected"]])
        for step in steps[1:]:
            ws.append(["", "", "", step["description"], "", step["expected"]])


def _write_api_sheet(wb, testcases):
    ws = wb.create_sheet(_API_SHEET)
    ws.append(_API_HEADERS)
    for tc in testcases:
        api = tc.get("api") or {}
        ws.append([tc["title"], tc["priority"], tc.get("precondition", ""),
                   api.get("endpoint", ""), api.get("method", ""),
                   api.get("request_headers", ""), api.get("request_body", ""),
                   api.get("expected_status", ""), api.get("expected_response", "")])


def _write_data_driven_sheet(wb, tc):
    ws = wb.create_sheet(_safe_sheet_name(tc["ref"], set(wb.sheetnames)))
    steps = tc["steps"] or [{"description": "", "expected": ""}]
    first = steps[0]
    ws.append([tc["title"], tc["priority"], tc.get("precondition", ""),
               first["description"], "Description", first["expected"]])
    for step in steps[1:]:
        ws.append(["", "", "", step["description"], "", step["expected"]])
    ws.append([_DATA_SEP_TEXT])
    all_keys = {k for v in tc["data_variants"] for k in v["values"]}
    variant_cols = sorted(all_keys - {"Expected"})
    ws.append(["", "", "", "", "Description"] + variant_cols + ["Expected"])
    for variant in tc["data_variants"]:
        row = ["", "", "", "", variant.get("label", "")]
        row += [variant["values"].get(col, "") for col in variant_cols]
        row.append(variant["values"].get("Expected", ""))
        ws.append(row)


def _selftest():
    testcases = [
        {"ref": "TC-1", "ac_refs": ["AC-1"], "title": "[TC-1] Happy path", "type": "Normal",
         "priority": "High", "precondition": "1. Logged in.",
         "steps": [{"description": "Click submit", "expected": "See success"}],
         "data_variants": [], "api": {}},
        {"ref": "TC-2", "ac_refs": ["AC-2"], "title": "[TC-2] Field variants", "type": "DataTable",
         "priority": "Medium", "precondition": "",
         "steps": [{"description": "Enter value", "expected": "Validated"}],
         "data_variants": [
             {"label": "empty", "values": {"Field": "", "Expected": "Error shown"}},
             {"label": "valid", "values": {"Field": "abc", "Expected": "Accepted"}},
         ], "api": {}},
        {"ref": "TC-3", "ac_refs": ["AC-3"], "title": "[TC-3] Login API", "type": "API",
         "priority": "Highest", "precondition": "",
         "steps": [], "data_variants": [],
         "api": {"endpoint": "/login", "method": "POST", "expected_status": "200"}},
    ]
    data = export_excel(testcases)
    wb = openpyxl.load_workbook(io.BytesIO(data))
    assert set(wb.sheetnames) == {"Normal_TestCases", "API_TestCases", "TC-2"}, wb.sheetnames
    normal = wb["Normal_TestCases"]
    assert [c.value for c in normal[1]] == _NORMAL_HEADERS
    assert normal["A2"].value == "[TC-1] Happy path"
    api_ws = wb["API_TestCases"]
    assert [c.value for c in api_ws[1]] == _API_HEADERS
    assert api_ws["A2"].value == "[TC-3] Login API"
    assert api_ws["D2"].value == "/login" and api_ws["E2"].value == "POST"
    dd = wb["TC-2"]
    assert dd["A1"].value == "[TC-2] Field variants"
    header_row = [c.value for c in dd[3]]
    assert header_row[4] == "Description" and header_row[-1] == "Expected", header_row
    assert dd[4][4].value == "empty" and (dd[4][5].value or "") == "" and dd[4][-1].value == "Error shown"

    try:
        export_excel([])
        raise AssertionError("expected ValueError for empty testcases list")
    except ValueError:
        pass

    return "ok"


if __name__ == "__main__":
    print(_selftest())
